import argparse
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

from supplier_common import ROOT, now_iso, parse_products_from_index, today_slug, write_products_to_index


MAPPING_PATH = ROOT / "data" / "estoque-scorsatto" / "mapeamento.json"


def normalize(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def normalize_size(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip().upper()


def normalize_quantity(value):
    try:
        quantity = int(float(value))
    except (TypeError, ValueError):
        return 0
    return max(quantity, 0)


def load_mapping():
    payload = json.loads(MAPPING_PATH.read_text(encoding="utf-8"))
    aliases = {}
    for item in payload["products"]:
        product_id = item["siteProductId"]
        for alias in item["aliases"]:
            key = normalize(alias)
            if key in aliases and aliases[key] != product_id:
                raise RuntimeError(f"Alias duplicado no mapeamento: {alias}")
            aliases[key] = product_id
    return payload, aliases


def load_inventory(path, sheet_name):
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Aba {sheet_name!r} nao encontrada. Abas: {workbook.sheetnames}")
    sheet = workbook[sheet_name]
    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        product = values[1] if len(values) > 1 else None
        size = values[2] if len(values) > 2 else None
        quantity = values[4] if len(values) > 4 else None
        if not product:
            continue
        rows.append({
            "row": row_number,
            "product": str(product).strip(),
            "size": normalize_size(size),
            "quantity": normalize_quantity(quantity),
        })
    return rows


def is_own_stock_product(product):
    return str(product.get("id", "")).startswith("sc-estoque-") or "estoque proprio" in normalize(product.get("supplierName"))


def main():
    parser = argparse.ArgumentParser(description="Sincroniza o estoque proprio do site com a planilha do Drive.")
    parser.add_argument("--xlsx", required=True, type=Path, help="Copia local do arquivo Financeiro Scorsatto.xlsx")
    parser.add_argument("--apply", action="store_true", help="Aplica as alteracoes seguras no index.html")
    parser.add_argument("--deactivate-missing", action="store_true", help="Retira da pronta entrega produtos mapeados ausentes da planilha")
    parser.add_argument("--report", type=Path, help="Caminho privado opcional para salvar o relatorio detalhado")
    args = parser.parse_args()

    mapping, aliases = load_mapping()
    inventory_rows = load_inventory(args.xlsx, mapping["sourceSheet"])
    products, html, span = parse_products_from_index()
    product_by_id = {product.get("id"): product for product in products}

    matched_rows = []
    pending_rows = []
    ignored_rows = []
    quantities = defaultdict(lambda: defaultdict(int))

    for row in inventory_rows:
        if normalize(row["size"]) == "materiais":
            ignored_rows.append({**row, "reason": "material interno"})
            continue
        product_id = aliases.get(normalize(row["product"]))
        if not product_id:
            pending_rows.append({**row, "reason": "produto sem mapeamento aprovado"})
            continue
        if not row["size"] or row["quantity"] <= 0:
            pending_rows.append({**row, "reason": "tamanho ou quantidade invalida"})
            continue
        quantities[product_id][row["size"]] += row["quantity"]
        matched_rows.append({**row, "siteProductId": product_id})

    missing_ids = sorted(product_id for product_id in quantities if product_id not in product_by_id)
    if missing_ids:
        raise RuntimeError(f"Produtos mapeados inexistentes no site: {missing_ids}")

    changes = []
    checked_at = today_slug()
    active_ids = set(quantities)
    mapped_ids = {item["siteProductId"] for item in mapping["products"]}

    for product_id, size_counts in sorted(quantities.items()):
        product = product_by_id[product_id]
        new_sizes = list(size_counts)
        new_stock = dict(size_counts)
        old_state = {
            "status": product.get("status"),
            "sizes": product.get("sizes") or [],
            "stock": product.get("stock") or {},
            "supplierName": product.get("supplierName"),
        }
        new_state = {"status": "disponivel", "sizes": new_sizes, "stock": new_stock, "supplierName": "Estoque proprio SCORSATTO"}
        new_tags = list(product.get("tags", []))
        for required_tag in ("disponiveis-agora", "estoque-proprio"):
            if required_tag not in new_tags:
                new_tags.append(required_tag)
        if old_state != new_state or product.get("tags") != new_tags:
            changes.append({"siteProductId": product_id, "name": product.get("name"), "from": old_state, "to": new_state})
        product["status"] = "disponivel"
        product["sizes"] = new_sizes
        product["stock"] = new_stock
        product["supplierName"] = "Estoque proprio SCORSATTO"
        product["tags"] = new_tags
        product["lastCheckedAt"] = checked_at

    deactivated = []
    if args.deactivate_missing:
        for product in products:
            product_id = product.get("id")
            if not is_own_stock_product(product) or product_id in active_ids or product_id not in mapped_ids:
                continue
            reason = "ausente na planilha"
            was_available = product.get("status") == "disponivel" or "disponiveis-agora" in product.get("tags", [])
            if was_available:
                deactivated.append({"siteProductId": product_id, "name": product.get("name"), "reason": reason})
                product["status"] = "vendido"
                product["stock"] = {}
                product["tags"] = [tag for tag in product.get("tags", []) if tag != "disponiveis-agora"]
                product["lastCheckedAt"] = checked_at

    if args.apply and (matched_rows or deactivated):
        write_products_to_index(products, html, span)

    report = {
        "generatedAt": now_iso(),
        "sourceSheet": mapping["sourceSheet"].strip(),
        "apply": args.apply,
        "deactivateMissing": args.deactivate_missing,
        "summary": {
            "inventoryRows": len(inventory_rows),
            "matchedRows": len(matched_rows),
            "pendingApprovalRows": len(pending_rows),
            "ignoredRows": len(ignored_rows),
            "changedProducts": len(changes),
            "deactivatedProducts": len(deactivated),
        },
        "matched": matched_rows,
        "pendingApproval": pending_rows,
        "ignored": ignored_rows,
        "changes": changes,
        "deactivated": deactivated,
    }
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"report": str(args.report) if args.report else None, **report["summary"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
